from __future__ import annotations

import argparse
import io
import re
import zipfile
from dataclasses import replace

import openpyxl
import numpy as np

from profile_processing import (
    EXCEL_POWER_EDGES,
    EXCEL_SPEED_EDGES,
    VesselValidationError,
    WrongReportTypeError,
    build_excel_data_sum,
    excel_overall_summary,
    fuel_consumption_summary,
    make_excel_profile,
    monthly_summary_excel,
    parse_all_report_files,
    parse_report_file,
    profile_segments_from_data_sum,
    sea_temperature_audit,
    validate_vessel_consistency,
)


def minimal_noon_workbook(drop_field: str | None = None, reverse_columns: bool = False) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Noon only"
    headers = [
        ("vessel", "Vessel", "Vessel"),
        ("voyage", "Voyage", "Voyage"),
        ("time", "Noon", "Time"),
        ("duration", "Data while Steaming", "Hours Propelling from Last Report to Noon"),
        ("speed", "Data while Steaming", "Ave. speed from Last Report to Noon"),
        ("me_output", "Data while Steaming", "M/E output (KW)"),
        ("data_sum_sea_temp", "Rolling", "Avg. Period(sec)"),
        ("sea_temp", "Sea Water", "temperature at noon"),
    ]
    headers = [item for item in headers if item[0] != drop_field]
    if reverse_columns:
        headers.reverse()
    for column, (_, parent, child) in enumerate(headers, 1):
        sheet.cell(1, column, parent)
        sheet.cell(2, column, child)
    sheet.append(["TEST", "V001", None, None, None, None])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def broken_dimension_noon_workbook() -> bytes:
    """Reproduce reports that declare A1:A1 while containing many populated cells."""
    source = io.BytesIO(minimal_noon_workbook())
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as input_archive, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as output_archive:
        for info in input_archive.infolist():
            content = input_archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                content = re.sub(
                    br'<dimension ref="[^"]+"\s*/>',
                    b'<dimension ref="A1"/>',
                    content,
                    count=1,
                )
            output_archive.writestr(info, content)
    return output.getvalue()


def run(workbook_path: str) -> None:
    with open(workbook_path, "rb") as handle:
        content = handle.read()
    reports, errors = parse_all_report_files(content)
    assert not errors, errors
    noon = reports["noon"]
    departure = reports["departure"]
    arrival = reports["arrival"]

    assert noon.sheet_name == "Noon", noon.sheet_name
    assert departure.sheet_name == "Dep", departure.sheet_name
    assert arrival.sheet_name == "Arr", arrival.sheet_name
    assert noon.mapping["sea_temp"].excel_column == "CN", noon.mapping["sea_temp"]
    assert noon.mapping["data_sum_sea_temp"].excel_column == "CM"
    assert departure.mapping["time"].excel_column == "AC", departure.mapping["time"]
    assert arrival.mapping["duration"].excel_column == "N", arrival.mapping["duration"]

    data_sum = build_excel_data_sum(noon, departure, arrival)
    segments = profile_segments_from_data_sum(data_sum)
    speed = make_excel_profile(
        segments, "speed_knots", EXCEL_SPEED_EDGES, "speed_included"
    )
    power = make_excel_profile(
        segments, "me_output_kw", EXCEL_POWER_EDGES, "power_included"
    )
    assert len(segments) == 265, len(segments)
    assert len(data_sum) == 405, len(data_sum)
    assert abs(speed.total_hours - 4_971.1) < 1e-8
    assert abs(power.total_hours - 5_009.9) < 1e-8
    assert abs(speed.percent.to_numpy().sum() - 100.0) < 1e-8
    assert abs(power.percent.to_numpy().sum() - 99.61077067406536) < 1e-8
    monthly = monthly_summary_excel(data_sum)
    assert len(monthly) == 12
    assert list(speed.percent.index) == [str(value) for value in range(7, 17)]
    assert list(speed.percent.columns) == [str(value) for value in range(9, 25)]
    assert list(power.percent.columns) == [str(value) for value in range(0, 23_000, 1_000)]

    overall = excel_overall_summary(data_sum)
    assert overall["year"] == 2024
    assert abs(overall["total_hours"] - 8_760) < 1e-8
    assert abs(overall["propelling_hours"] - 5_161.9) < 1e-8
    assert abs(overall["avg_speed_knots"] - 14.130722433460074) < 1e-8
    assert abs(overall["avg_sea_temp_excel"] - 24.132075471698112) < 1e-8
    assert abs(overall["avg_actual_sea_temp_c"] - 24.132075471698112) < 1e-8
    temperature = sea_temperature_audit(data_sum, noon)
    assert temperature["valid"]
    assert temperature["source_column"] == "Internal Data_sum G"
    assert temperature["upstream_column"] == "CN"
    assert temperature["numeric_count"] == 265
    assert temperature["zero_count"] == 3
    assert abs(temperature["sum_value"] - 6_395.0) < 1e-8
    assert abs(temperature["average_value"] - 24.132075471698112) < 1e-8

    expected_monthly_sea_temp = np.array(
        [
            26, 23.791666666666668, 24.142857142857142, 23.85,
            24.764705882352942, 24.818181818181817, 27.166666666666668,
            24.333333333333332, 24.434782608695652, 25.333333333333332,
            23.555555555555557, 24.181818181818183,
        ]
    )
    assert np.allclose(
        monthly["avg_sea_temp_excel"].head(12).to_numpy(),
        expected_monthly_sea_temp,
        atol=1e-12,
    )
    assert np.allclose(
        monthly["available_hours"].head(12).to_numpy(),
        [0, 696, 720, 703.5, 721, 693.5, 707.1, 693, 670.6, 720, 729.8, 628.3],
        atol=1e-8,
    )
    assert np.allclose(
        monthly["propelling_hours"].head(12).to_numpy(),
        [24, 451.9, 445, 404, 367.2, 382.8, 364.2, 373.9, 507.4, 525.4, 529.8, 407.6],
        atol=1e-8,
    )
    assert np.allclose(
        monthly["avg_speed_knots"].head(12).to_numpy(),
        [11.3, 13.366666666666665, 13.542857142857144, 15.285,
         15.747058823529407, 15.331818181818182, 15.488888888888887, 13.4,
         15.30521739130435, 14.904166666666667, 11.733333333333336,
         13.231818181818182],
        atol=1e-12,
    )
    fuel = fuel_consumption_summary(noon, arrival)
    assert abs(fuel["noon_vlsfo"] - 8_947.2) < 1e-8
    assert abs(fuel["arrival_vlsfo"] - 1_043.5) < 1e-8
    assert fuel["noon_mgo"] == 0
    assert fuel["arrival_mgo"] == 0

    reference = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    expected_speed = np.array(
        [
            [reference["Profile"].cell(row, column).value or 0 for column in range(3, 19)]
            for row in range(3, 13)
        ],
        dtype=float,
    ) * 100
    expected_power = np.array(
        [
            [
                reference["Profile (ME Output)"].cell(row, column).value or 0
                for column in range(3, 26)
            ]
            for row in range(3, 13)
        ],
        dtype=float,
    ) * 100
    assert np.allclose(speed.percent.to_numpy(), expected_speed, atol=1e-12)
    assert np.allclose(power.percent.to_numpy(), expected_power, atol=1e-12)

    try:
        parse_report_file(minimal_noon_workbook(), "departure")
    except WrongReportTypeError:
        pass
    else:
        raise AssertionError("A Noon report uploaded as Departure was not rejected")

    reordered = parse_report_file(minimal_noon_workbook(reverse_columns=True), "noon")
    assert not reordered.missing, reordered.missing
    missing = parse_report_file(minimal_noon_workbook(drop_field="speed"), "noon")
    assert missing.missing == ["speed"], missing.missing
    missing_temperature = parse_report_file(
        minimal_noon_workbook(drop_field="sea_temp"), "noon"
    )
    assert missing_temperature.missing == ["sea_temp"], missing_temperature.missing
    no_rolling_period = parse_report_file(
        minimal_noon_workbook(drop_field="data_sum_sea_temp"), "noon"
    )
    assert not no_rolling_period.missing, no_rolling_period.missing
    recovered = parse_report_file(broken_dimension_noon_workbook(), "noon")
    assert recovered.confidence == 1.0
    assert len(recovered.data) == 1
    assert any("recovered automatically" in warning for warning in recovered.warnings)

    assert validate_vessel_consistency(
        {"noon": noon, "departure": departure, "arrival": arrival}
    ) == "NYK FUTAGO"

    wrong_arrival_data = arrival.data.copy()
    wrong_arrival_data["vessel"] = "OTHER VESSEL"
    wrong_arrival = replace(arrival, data=wrong_arrival_data)
    try:
        validate_vessel_consistency(
            {"noon": noon, "departure": departure, "arrival": wrong_arrival}
        )
    except VesselValidationError as exc:
        assert "Vessel mismatch" in str(exc)
        assert "NYK FUTAGO" in str(exc) and "OTHER VESSEL" in str(exc)
    else:
        raise AssertionError("A different-vessel Arrival report was not rejected")

    mixed_noon_data = noon.data.copy()
    mixed_noon_data.loc[mixed_noon_data.index[0], "vessel"] = "OTHER VESSEL"
    mixed_noon = replace(noon, data=mixed_noon_data)
    try:
        validate_vessel_consistency(
            {"noon": mixed_noon, "departure": departure, "arrival": arrival}
        )
    except VesselValidationError as exc:
        assert "multiple vessel names" in str(exc)
    else:
        raise AssertionError("A mixed-vessel Noon report was not rejected")

    blank_departure_data = departure.data.copy()
    blank_departure_data["vessel"] = None
    blank_departure = replace(departure, data=blank_departure_data)
    try:
        validate_vessel_consistency(
            {"noon": noon, "departure": blank_departure, "arrival": arrival}
        )
    except VesselValidationError as exc:
        assert "contains no vessel name" in str(exc)
    else:
        raise AssertionError("A report with no vessel name was not rejected")

    print(
        f"PASS: {len(data_sum)} Data_sum rows; monthly table and both profile matrices match Excel"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    run(parser.parse_args().workbook)
