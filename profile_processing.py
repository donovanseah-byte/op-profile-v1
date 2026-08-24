from __future__ import annotations

import io
import math
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any, Iterable

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter


REPORT_NAMES = {"noon": "Noon", "departure": "Departure", "arrival": "Arrival"}

EXCEL_DRAFT_EDGES = np.arange(7.0, 18.0, 1.0)
EXCEL_SPEED_EDGES = np.arange(9.0, 26.0, 1.0)
EXCEL_POWER_EDGES = np.arange(0.0, 24_000.0, 1_000.0)


FIELD_ALIASES: dict[str, dict[str, list[str]]] = {
    "noon": {
        "imo": ["imo number", "imo no"],
        "vessel": ["vessel"],
        "voyage": ["voyage"],
        "time": ["noon time"],
        "duration": [
            "data while steaming hours propelling from last report to noon",
            "hours propelling from last report to noon",
        ],
        "speed": [
            "data while steaming ave speed from last report to noon",
            "ave speed from last report to noon",
            "average speed from last report to noon",
        ],
        "me_output": ["data while steaming me output kw", "me output kw"],
        "data_sum_sea_temp": ["rolling avg period sec", "rolling average period sec"],
        "sea_temp": [
            "sea water temperature at noon",
            "seawater temperature at noon",
            "sea water temp at noon",
            "seawater temp at noon",
        ],
        "foc_vlsfo": [
            "consumption while steaming from last report to noon me foc vlsfo"
        ],
        "foc_mgo": [
            "consumption while steaming from last report to noon me foc mdo mgo"
        ],
    },
    "departure": {
        "imo": ["imo number", "imo no"],
        "vessel": ["vessel"],
        "voyage": ["voyage"],
        "time": [
            "sop start of sea passage detail time",
            "start of sea passage detail time",
            "sop detail time",
        ],
        "draft_mid": ["departure condition draft midship", "departure draft midship"],
        "departure_time": [
            "harbour operation detail time of departure",
            "time of departure",
        ],
    },
    "arrival": {
        "imo": ["imo number", "imo no"],
        "vessel": ["vessel"],
        "voyage": ["voyage"],
        "time": [
            "eop end of sea passage detail time",
            "end of sea passage detail time",
            "eop detail time",
        ],
        "duration": [
            "data while steaming hours propelling from last report to eop",
            "hours propelling from last report to eop",
        ],
        "speed": [
            "data while steaming ave speed from last report to eop",
            "ave speed from last report to eop",
            "average speed from last report to eop",
        ],
        "me_output": ["data while steaming me output kw", "me output kw"],
        "draft_mid": ["arrival condition draft midship", "arrival draft midship"],
        "foc_vlsfo": [
            "consumption while steaming from last report to eop me vlsfo"
        ],
        "foc_mgo": [
            "consumption while steaming from last report to eop me mdo mgo"
        ],
    },
}


REQUIRED_FIELDS = {
    "noon": [
        "vessel", "voyage", "time", "duration", "speed", "me_output", "sea_temp"
    ],
    "departure": ["vessel", "voyage", "time", "draft_mid"],
    "arrival": ["vessel", "voyage", "time", "duration", "speed", "me_output", "draft_mid"],
}


FIELD_TITLES = {
    "imo": "IMO number",
    "vessel": "Vessel",
    "voyage": "Voyage",
    "time": "Report time",
    "duration": "Hours propelling",
    "speed": "Average speed",
    "me_output": "M/E output (kW)",
    "data_sum_sea_temp": "Rolling > Avg. Period(sec) (diagnostic only)",
    "sea_temp": "Sea water temperature at noon (Data_sum source)",
    "draft_mid": "Draft midship",
    "departure_time": "Time of departure",
    "foc_vlsfo": "M/E VLSFO consumption (MT)",
    "foc_mgo": "M/E MDO/MGO consumption (MT)",
}


class ReportError(ValueError):
    pass


class WrongReportTypeError(ReportError):
    def __init__(self, expected: str, detected: str, sheet_name: str, confidence: float):
        self.expected = expected
        self.detected = detected
        self.sheet_name = sheet_name
        self.confidence = confidence
        super().__init__(
            f"Wrong report type detected. {REPORT_NAMES[expected]} expected, but this file "
            f"matches {REPORT_NAMES[detected]} ({confidence:.0%}) on sheet '{sheet_name}'."
        )


class UnknownReportError(ReportError):
    pass


class VesselValidationError(ReportError):
    pass


@dataclass
class HeaderColumn:
    index: int
    excel_column: str
    parent: str
    child: str
    label: str
    normalized: str


@dataclass
class SheetInspection:
    sheet_name: str
    headers: list[HeaderColumn]
    mappings: dict[str, dict[str, HeaderColumn]] = field(default_factory=dict)
    scores: dict[str, float] = field(default_factory=dict)


@dataclass
class ParsedReport:
    report_type: str
    sheet_name: str
    confidence: float
    data: pd.DataFrame
    mapping: dict[str, HeaderColumn]
    missing: list[str]
    warnings: list[str]


@dataclass
class ProfileResult:
    hours: pd.DataFrame
    percent: pd.DataFrame
    total_hours: float
    eligible_hours: float
    coverage: float
    draft_edges: np.ndarray
    x_edges: np.ndarray


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("M/E", "ME").replace("m/e", "me")
    text = text.replace("&", " and ")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).lower()
    return re.sub(r"\s+", " ", text).strip()


def _sheet_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[HeaderColumn]:
    merged_parent_by_col: dict[int, Any] = {}
    if hasattr(ws, "merged_cells"):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= 1 <= merged.max_row:
                value = ws.cell(merged.min_row, merged.min_col).value
                for col in range(merged.min_col, merged.max_col + 1):
                    merged_parent_by_col[col] = value

    headers: list[HeaderColumn] = []
    last_parent: Any = None
    for col in range(1, ws.max_column + 1):
        raw_parent = merged_parent_by_col.get(col, ws.cell(1, col).value)
        if raw_parent not in (None, ""):
            last_parent = raw_parent
        parent = raw_parent if raw_parent not in (None, "") else last_parent
        child = ws.cell(2, col).value
        parent_text = "" if parent is None else str(parent).strip()
        child_text = "" if child is None else str(child).strip()
        if normalize_header(parent_text) == normalize_header(child_text):
            label = parent_text
        else:
            label = " > ".join(part for part in (parent_text, child_text) if part)
        headers.append(
            HeaderColumn(
                index=col,
                excel_column=get_column_letter(col),
                parent=parent_text,
                child=child_text,
                label=label or f"Column {get_column_letter(col)}",
                normalized=normalize_header(label),
            )
        )
    return headers


def _has_broken_worksheet_dimensions(file_bytes: bytes, workbook) -> bool:
    """Detect exports whose XML says A1:A1 even though later cells exist.

    OpenPyXL's read-only mode trusts the declared worksheet dimension. Some reporting
    systems export valid rows and columns but leave that dimension as A1, causing only
    the first cell to be visible. The raw XML check avoids slowing down normal files.
    """
    suspicious = any(
        ws.calculate_dimension(force=True) in {"A1", "A1:A1"} for ws in workbook.worksheets
    )
    if not suspicious:
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            for name in archive.namelist():
                if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                    continue
                xml = archive.read(name)
                declared_as_a1 = re.search(br'<dimension\s+ref="A1(?::A1)?"', xml)
                later_cell_exists = re.search(br'<c\s+r="(?!A1")[A-Z]+[0-9]+"', xml)
                if declared_as_a1 and later_cell_exists:
                    return True
    except (zipfile.BadZipFile, KeyError):
        return False
    return False


def _load_excel_workbook(file_bytes: bytes):
    """Load quickly when possible, but recover malformed report-system exports."""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        if _has_broken_worksheet_dimensions(file_bytes, workbook):
            workbook.close()
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_bytes), data_only=True, read_only=False
            )
            return workbook, True
        return workbook, False
    except Exception as exc:
        raise ReportError(f"The uploaded file could not be read as an Excel workbook: {exc}") from exc


def _alias_match_score(header: str, alias: str) -> float:
    if not header or not alias:
        return 0.0
    if header == alias:
        return 1.0
    if header.endswith(alias) or alias in header:
        return 0.96
    header_tokens = set(header.split())
    alias_tokens = set(alias.split())
    if not alias_tokens:
        return 0.0
    coverage = len(header_tokens & alias_tokens) / len(alias_tokens)
    precision = len(header_tokens & alias_tokens) / max(len(header_tokens), 1)
    if coverage == 1.0:
        return 0.82 + min(precision, 0.14)
    return 0.0


def _map_fields(headers: list[HeaderColumn], report_type: str) -> dict[str, HeaderColumn]:
    result: dict[str, HeaderColumn] = {}
    used_columns: set[int] = set()
    for field_name, aliases in FIELD_ALIASES[report_type].items():
        candidates: list[tuple[float, int, HeaderColumn]] = []
        for header in headers:
            if header.index in used_columns:
                continue
            best = max(
                (_alias_match_score(header.normalized, normalize_header(alias)) for alias in aliases),
                default=0.0,
            )
            if best >= 0.82:
                candidates.append((best, len(header.normalized), header))
        if candidates:
            candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
            selected = candidates[0][2]
            if field_name == "sea_temp" and any(
                token in selected.normalized for token in ("rolling", "average period", "seconds", " sec")
            ):
                continue
            result[field_name] = selected
            used_columns.add(selected.index)
    return result


def inspect_workbook(file_bytes: bytes) -> list[SheetInspection]:
    workbook, _ = _load_excel_workbook(file_bytes)
    return _inspect_loaded_workbook(workbook)


def _inspect_loaded_workbook(workbook) -> list[SheetInspection]:
    inspections: list[SheetInspection] = []
    for ws in workbook.worksheets:
        headers = _sheet_headers(ws)
        mappings: dict[str, dict[str, HeaderColumn]] = {}
        scores: dict[str, float] = {}
        for report_type in REPORT_NAMES:
            mapping = _map_fields(headers, report_type)
            required = REQUIRED_FIELDS[report_type]
            mappings[report_type] = mapping
            scores[report_type] = sum(field in mapping for field in required) / len(required)
        inspections.append(SheetInspection(ws.title, headers, mappings, scores))
    return inspections


def _extract_rows(workbook, sheet_name: str, mapping: dict[str, HeaderColumn]) -> pd.DataFrame:
    ws = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        record = {
            field: values[header.index - 1] if header.index - 1 < len(values) else None
            for field, header in mapping.items()
        }
        record["_source_row"] = row_number
        if any(value not in (None, "") for key, value in record.items() if not key.startswith("_")):
            rows.append(record)
    return pd.DataFrame(rows)


def _parse_loaded_report(
    workbook,
    inspections: list[SheetInspection],
    expected_type: str,
    recovered_dimensions: bool = False,
) -> ParsedReport:
    if expected_type not in REPORT_NAMES:
        raise ValueError(f"Unsupported report type: {expected_type}")
    expected_sheet = max(inspections, key=lambda item: item.scores[expected_type])
    expected_score = expected_sheet.scores[expected_type]

    global_options = [
        (inspection.scores[report_type], report_type, inspection)
        for inspection in inspections
        for report_type in REPORT_NAMES
    ]
    global_score, global_type, global_sheet = max(global_options, key=lambda item: item[0])
    if (
        global_type != expected_type
        and global_score >= 0.60
        and expected_score < 0.75
        and global_score >= expected_score + 0.15
    ):
        raise WrongReportTypeError(expected_type, global_type, global_sheet.sheet_name, global_score)
    if expected_score < 0.45:
        raise UnknownReportError(
            f"Could not identify a {REPORT_NAMES[expected_type]} report. "
            "Check that the file contains the original two header rows."
        )

    mapping = expected_sheet.mappings[expected_type]
    missing = [field for field in REQUIRED_FIELDS[expected_type] if field not in mapping]
    warnings: list[str] = []
    if recovered_dimensions:
        warnings.append(
            "The workbook had invalid worksheet-range metadata; the full sheet was recovered automatically."
        )
    if missing:
        warnings.append(
            "Missing required headers: " + ", ".join(FIELD_TITLES.get(field, field) for field in missing)
        )
    data = _extract_rows(workbook, expected_sheet.sheet_name, mapping)
    if data.empty:
        warnings.append("The matched sheet has no populated data rows beneath its headers.")
    return ParsedReport(
        report_type=expected_type,
        sheet_name=expected_sheet.sheet_name,
        confidence=expected_score,
        data=data,
        mapping=mapping,
        missing=missing,
        warnings=warnings,
    )


def parse_report_file(file_bytes: bytes, expected_type: str) -> ParsedReport:
    if expected_type not in REPORT_NAMES:
        raise ValueError(f"Unsupported report type: {expected_type}")
    workbook, recovered_dimensions = _load_excel_workbook(file_bytes)
    inspections = _inspect_loaded_workbook(workbook)
    return _parse_loaded_report(
        workbook, inspections, expected_type, recovered_dimensions=recovered_dimensions
    )


def parse_all_report_files(file_bytes: bytes) -> tuple[dict[str, ParsedReport], dict[str, str]]:
    """Parse every expected report type with one workbook load.

    This is especially useful when the same combined workbook is placed in all three upload slots.
    """
    try:
        workbook, recovered_dimensions = _load_excel_workbook(file_bytes)
    except ReportError as exc:
        message = str(exc)
        return {}, {report_type: message for report_type in REPORT_NAMES}
    inspections = _inspect_loaded_workbook(workbook)
    reports: dict[str, ParsedReport] = {}
    errors: dict[str, str] = {}
    for report_type in REPORT_NAMES:
        try:
            reports[report_type] = _parse_loaded_report(
                workbook,
                inspections,
                report_type,
                recovered_dimensions=recovered_dimensions,
            )
        except ReportError as exc:
            errors[report_type] = str(exc)
    return reports, errors


def _clean_key(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def _to_datetime(value: Any) -> pd.Timestamp:
    if value is None or value == "":
        return pd.NaT
    try:
        return pd.Timestamp(value)
    except Exception:
        return pd.to_datetime(value, errors="coerce", dayfirst=True)


def duration_to_hours(value: Any) -> float:
    if value is None or value == "":
        return np.nan
    if isinstance(value, pd.Timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float, np.number)) and not isinstance(value, bool):
        number = float(value)
        if not math.isfinite(number):
            return np.nan
        # Excel stores elapsed time as a fraction of a day. Large values are usually already hours.
        return number * 24 if 0 <= number < 4 else number
    text = str(value).strip()
    match = re.fullmatch(r"(-?\d+):([0-5]?\d)(?::([0-5]?\d))?", text)
    if match:
        hours, minutes, seconds = match.groups()
        sign = -1 if int(hours) < 0 else 1
        return int(hours) + sign * int(minutes) / 60 + sign * int(seconds or 0) / 3600
    parsed = pd.to_timedelta(text, errors="coerce")
    if pd.isna(parsed):
        return np.nan
    return parsed.total_seconds() / 3600


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _standardize(report: ParsedReport) -> pd.DataFrame:
    frame = report.data.copy()
    for column in ("vessel", "voyage"):
        if column not in frame:
            frame[column] = ""
        frame[column] = frame[column].map(_clean_key)
    if "time" in frame:
        frame["timestamp"] = frame["time"].map(_to_datetime)
    else:
        frame["timestamp"] = pd.NaT
    if "duration" in frame:
        frame["duration_hours"] = frame["duration"].map(duration_to_hours)
    for original, new_name in (
        ("speed", "speed_knots"),
        ("me_output", "me_output_kw"),
        ("data_sum_sea_temp", "data_sum_sea_temp_value"),
        ("sea_temp", "sea_temp_c"),
        ("draft_mid", "draft_m"),
    ):
        frame[new_name] = _numeric(frame[original]) if original in frame else np.nan
    return frame


def _lookup_draft(
    timestamp: pd.Timestamp,
    vessel: str,
    voyage: str,
    departures: pd.DataFrame,
    arrivals: pd.DataFrame,
    method: str,
) -> tuple[float, str]:
    if pd.isna(timestamp):
        return np.nan, "No report time"

    def matching(frame: pd.DataFrame) -> pd.DataFrame:
        candidate = frame[(frame["vessel"] == vessel) & (frame["voyage"] == voyage)]
        if candidate.empty and vessel:
            candidate = frame[frame["vessel"] == vessel]
        return candidate

    dep = matching(departures)
    arr = matching(arrivals)
    dep = dep[(dep["timestamp"] <= timestamp) & dep["draft_m"].notna()].sort_values("timestamp")
    arr = arr[(arr["timestamp"] >= timestamp) & arr["draft_m"].notna()].sort_values("timestamp")
    dep_row = dep.iloc[-1] if not dep.empty else None
    arr_row = arr.iloc[0] if not arr.empty else None

    if dep_row is not None and arr_row is not None:
        if method == "interpolate" and arr_row["timestamp"] > dep_row["timestamp"]:
            ratio = (timestamp - dep_row["timestamp"]) / (arr_row["timestamp"] - dep_row["timestamp"])
            ratio = float(np.clip(ratio, 0.0, 1.0))
            value = dep_row["draft_m"] + ratio * (arr_row["draft_m"] - dep_row["draft_m"])
            return float(value), "Interpolated departure→arrival"
        return float(dep_row["draft_m"]), "Departure draft"
    if dep_row is not None:
        return float(dep_row["draft_m"]), "Departure draft (arrival unavailable)"
    if arr_row is not None:
        return float(arr_row["draft_m"]), "Arrival draft (departure unavailable)"
    return np.nan, "No matching departure/arrival draft"


def _reason(row: pd.Series, profile: str) -> str:
    reasons: list[str] = []
    if pd.isna(row["timestamp"]):
        reasons.append("missing time")
    if pd.isna(row["duration_hours"]) or not (0 < row["duration_hours"] <= 72):
        reasons.append("duration outside 0–72 h")
    if pd.isna(row["draft_m"]) or not (0 < row["draft_m"] <= 30):
        reasons.append("draft outside 0–30 m")
    if profile == "speed" and (pd.isna(row["speed_knots"]) or not (0 < row["speed_knots"] <= 40)):
        reasons.append("speed outside 0–40 kn")
    if profile == "power" and (
        pd.isna(row["me_output_kw"]) or not (0 < row["me_output_kw"] <= 100_000)
    ):
        reasons.append("M/E output outside 0–100,000 kW")
    return "; ".join(reasons)


def build_segments(
    noon: ParsedReport,
    departure: ParsedReport,
    arrival: ParsedReport,
    include_arrival: bool = True,
    draft_method: str = "interpolate",
) -> pd.DataFrame:
    noon_df = _standardize(noon)
    dep_df = _standardize(departure)
    arr_df = _standardize(arrival)

    records: list[dict[str, Any]] = []
    for _, row in noon_df.iterrows():
        draft, draft_source = _lookup_draft(
            row["timestamp"], row["vessel"], row["voyage"], dep_df, arr_df, draft_method
        )
        duration = row.get("duration_hours", np.nan)
        timestamp = row["timestamp"]
        records.append(
            {
                "source": "Noon",
                "source_row": row.get("_source_row"),
                "vessel": row["vessel"],
                "voyage": row["voyage"],
                "segment_end": timestamp,
                "segment_start": timestamp - pd.Timedelta(hours=duration)
                if pd.notna(timestamp) and pd.notna(duration)
                else pd.NaT,
                "duration_hours": duration,
                "speed_knots": row.get("speed_knots", np.nan),
                "me_output_kw": row.get("me_output_kw", np.nan),
                "sea_temp_c": row.get("sea_temp_c", np.nan),
                "draft_m": draft,
                "draft_source": draft_source,
            }
        )

    if include_arrival:
        for _, row in arr_df.iterrows():
            duration = row.get("duration_hours", np.nan)
            timestamp = row["timestamp"]
            records.append(
                {
                    "source": "Arrival",
                    "source_row": row.get("_source_row"),
                    "vessel": row["vessel"],
                    "voyage": row["voyage"],
                    "segment_end": timestamp,
                    "segment_start": timestamp - pd.Timedelta(hours=duration)
                    if pd.notna(timestamp) and pd.notna(duration)
                    else pd.NaT,
                    "duration_hours": duration,
                    "speed_knots": row.get("speed_knots", np.nan),
                    "me_output_kw": row.get("me_output_kw", np.nan),
                    "sea_temp_c": np.nan,
                    "draft_m": row.get("draft_m", np.nan),
                    "draft_source": "Arrival draft",
                }
            )

    columns = [
        "source", "source_row", "vessel", "voyage", "segment_start", "segment_end",
        "duration_hours", "speed_knots", "me_output_kw", "sea_temp_c", "draft_m", "draft_source",
    ]
    segments = pd.DataFrame(records, columns=columns)
    if segments.empty:
        segments["speed_exclusion"] = []
        segments["power_exclusion"] = []
        segments["speed_included"] = []
        segments["power_included"] = []
        return segments
    segments = segments.drop_duplicates(
        subset=["source", "vessel", "voyage", "segment_end", "duration_hours"], keep="first"
    ).reset_index(drop=True)
    segments["timestamp"] = segments["segment_end"]
    segments["speed_exclusion"] = segments.apply(_reason, axis=1, profile="speed")
    segments["power_exclusion"] = segments.apply(_reason, axis=1, profile="power")
    segments["speed_included"] = segments["speed_exclusion"].eq("")
    segments["power_included"] = segments["power_exclusion"].eq("")
    return segments


def _excel_exclusion_reason(row: pd.Series, profile: str) -> str:
    reasons: list[str] = []
    if pd.isna(row["timestamp"]):
        reasons.append("missing report time")
    if pd.isna(row["duration_hours"]) or row["duration_hours"] <= 0:
        reasons.append("duration is blank or zero")
    if pd.isna(row["draft_m"]) or row["draft_m"] < 7:
        reasons.append("draft below 7 m or unavailable")
    if profile == "speed" and (pd.isna(row["speed_knots"]) or row["speed_knots"] < 9):
        reasons.append("speed below 9 kn or unavailable")
    if profile == "power" and (pd.isna(row["me_output_kw"]) or row["me_output_kw"] < 0):
        reasons.append("M/E output unavailable or negative")
    return "; ".join(reasons)


def build_excel_data_sum(
    noon: ParsedReport,
    departure: ParsedReport,
    arrival: ParsedReport,
) -> pd.DataFrame:
    """Build the internal Data_sum table before any profile or graph calculation.

    Departure sets the active midship draft at Time of Departure. Arrival/EOP
    resets the active draft to zero. Noon rows inherit the latest event state.
    Like the workbook, Departure and Arrival rows contribute timestamps but zero
    duration; Data_sum Sea Water Temp is populated only on Noon rows from the
    title-mapped physical Sea Water temperature at noon field.
    """
    noon_df = _standardize(noon)
    dep_df = _standardize(departure)
    arr_df = _standardize(arrival)

    if "departure_time" in dep_df:
        dep_df["event_time"] = dep_df["departure_time"].map(_to_datetime)
        dep_df["event_time"] = dep_df["event_time"].fillna(dep_df["timestamp"])
    else:
        dep_df["event_time"] = dep_df["timestamp"]
    arr_df["event_time"] = arr_df["timestamp"]

    dep_events = dep_df[["vessel", "event_time", "draft_m"]].copy()
    dep_events["event_type"] = "Departure"
    arr_events = arr_df[["vessel", "event_time"]].copy()
    arr_events["draft_m"] = 0.0
    arr_events["event_type"] = "Arrival"
    events = pd.concat([dep_events, arr_events], ignore_index=True)
    events = events[events["event_time"].notna()].sort_values(
        ["vessel", "event_time", "event_type"]
    )

    records: list[dict[str, Any]] = []
    for _, row in noon_df.iterrows():
        prior = events[
            (events["vessel"] == row["vessel"]) & (events["event_time"] <= row["timestamp"])
        ]
        if prior.empty:
            draft = np.nan
            draft_source = "No earlier Departure/Arrival state"
        else:
            event = prior.iloc[-1]
            draft = event["draft_m"]
            draft_source = f"Latest {event['event_type']} state at {event['event_time']}"
        duration = row.get("duration_hours", np.nan)
        timestamp = row["timestamp"]
        records.append(
            {
                "source": "Noon",
                "source_row": row.get("_source_row"),
                "vessel": row["vessel"],
                "voyage": row["voyage"],
                "timestamp": timestamp,
                "duration_hours": duration,
                "speed_knots": row.get("speed_knots", np.nan),
                "draft_m": draft,
                "data_sum_sea_temp": row.get("sea_temp_c", np.nan),
                "actual_sea_temp_c": row.get("sea_temp_c", np.nan),
                "me_output_kw": row.get("me_output_kw", np.nan),
                "draft_source": draft_source,
            }
        )

    for _, row in dep_df.iterrows():
        records.append(
            {
                "source": "Departure",
                "source_row": row.get("_source_row"),
                "vessel": row["vessel"],
                "voyage": row["voyage"],
                "timestamp": row["event_time"],
                "duration_hours": 0.0,
                "speed_knots": np.nan,
                "draft_m": row.get("draft_m", np.nan),
                "data_sum_sea_temp": np.nan,
                "actual_sea_temp_c": np.nan,
                "me_output_kw": np.nan,
                "draft_source": "Departure draft",
            }
        )

    for _, row in arr_df.iterrows():
        records.append(
            {
                "source": "Arrival",
                "source_row": row.get("_source_row"),
                "vessel": row["vessel"],
                "voyage": row["voyage"],
                "timestamp": row["event_time"],
                "duration_hours": 0.0,
                "speed_knots": np.nan,
                "draft_m": 0.0,
                "data_sum_sea_temp": np.nan,
                "actual_sea_temp_c": np.nan,
                "me_output_kw": row.get("me_output_kw", np.nan),
                "draft_source": "Arrival reset",
            }
        )

    data_sum = pd.DataFrame(records)
    if data_sum.empty:
        return data_sum
    data_sum = data_sum[data_sum["timestamp"].notna()].copy()
    priority = {"Arrival": 0, "Departure": 1, "Noon": 2}
    data_sum["_event_priority"] = data_sum["source"].map(priority).fillna(9)
    data_sum = data_sum.sort_values(
        ["timestamp", "_event_priority", "source_row"], kind="stable"
    ).drop(columns="_event_priority").reset_index(drop=True)
    data_sum.insert(0, "data_sum_row", np.arange(2, len(data_sum) + 2))
    data_sum["year"] = data_sum["timestamp"].dt.year
    data_sum["month"] = data_sum["timestamp"].dt.month
    data_sum["day"] = data_sum["timestamp"].dt.day
    data_sum["hour"] = data_sum["timestamp"].dt.hour
    data_sum["minute"] = data_sum["timestamp"].dt.minute
    data_sum["duration_days"] = data_sum["duration_hours"].fillna(0) / 24
    return data_sum


def profile_segments_from_data_sum(data_sum: pd.DataFrame) -> pd.DataFrame:
    """Select the Noon rows used by the profile matrices from internal Data_sum."""
    segments = data_sum[data_sum["source"].eq("Noon")].copy()
    if segments.empty:
        for column in ("speed_exclusion", "power_exclusion", "speed_included", "power_included"):
            segments[column] = []
        return segments
    segments["segment_end"] = segments["timestamp"]
    segments["segment_start"] = segments["timestamp"] - pd.to_timedelta(
        segments["duration_hours"], unit="h"
    )
    segments["speed_exclusion"] = segments.apply(
        _excel_exclusion_reason, axis=1, profile="speed"
    )
    segments["power_exclusion"] = segments.apply(
        _excel_exclusion_reason, axis=1, profile="power"
    )
    segments["speed_included"] = segments["speed_exclusion"].eq("")
    segments["power_included"] = segments["power_exclusion"].eq("")
    return segments


def build_excel_profile_segments(
    noon: ParsedReport,
    departure: ParsedReport,
    arrival: ParsedReport,
) -> pd.DataFrame:
    """Compatibility wrapper returning profile rows from the internal Data_sum."""
    return profile_segments_from_data_sum(build_excel_data_sum(noon, departure, arrival))


def _edges(values: pd.Series, width: float) -> np.ndarray:
    clean = pd.to_numeric(values, errors="coerce").dropna()
    if clean.empty:
        return np.array([])
    lower = math.floor(clean.min() / width) * width
    upper = math.ceil(clean.max() / width) * width
    if math.isclose(upper, clean.max()):
        upper += width
    if upper <= lower:
        upper = lower + width
    count = int(round((upper - lower) / width))
    return lower + np.arange(count + 1) * width


def _format_bin(value: float) -> str:
    return f"{value:g}"


def make_profile(
    segments: pd.DataFrame,
    x_column: str,
    x_width: float,
    draft_width: float,
    included_column: str,
) -> ProfileResult:
    eligible = segments[segments[included_column]].copy()
    total_duration_rows = segments[
        segments["duration_hours"].notna()
        & segments["duration_hours"].gt(0)
        & segments["duration_hours"].le(72)
    ]
    eligible_hours = float(total_duration_rows["duration_hours"].sum())
    if eligible.empty:
        return ProfileResult(
            pd.DataFrame(), pd.DataFrame(), 0.0, eligible_hours, 0.0, np.array([]), np.array([])
        )
    draft_edges = _edges(eligible["draft_m"], draft_width)
    x_edges = _edges(eligible[x_column], x_width)
    histogram, _, _ = np.histogram2d(
        eligible["draft_m"],
        eligible[x_column],
        bins=[draft_edges, x_edges],
        weights=eligible["duration_hours"],
    )
    total_hours = float(histogram.sum())
    percent = histogram / total_hours * 100 if total_hours else histogram
    draft_labels = [_format_bin(value) for value in draft_edges[:-1]]
    x_labels = [_format_bin(value) for value in x_edges[:-1]]
    hours_df = pd.DataFrame(histogram, index=draft_labels, columns=x_labels)
    percent_df = pd.DataFrame(percent, index=draft_labels, columns=x_labels)
    coverage = total_hours / eligible_hours if eligible_hours else 0.0
    return ProfileResult(hours_df, percent_df, total_hours, eligible_hours, coverage, draft_edges, x_edges)


def make_excel_profile(
    segments: pd.DataFrame,
    x_column: str,
    x_edges: np.ndarray,
    included_column: str,
) -> ProfileResult:
    """Calculate a fixed-bin matrix using the same denominator as the Excel formulas."""
    eligible = segments[segments[included_column]].copy()
    denominator_hours = float(eligible["duration_hours"].sum())
    draft_edges = EXCEL_DRAFT_EDGES.copy()
    histogram, _, _ = np.histogram2d(
        eligible["draft_m"],
        eligible[x_column],
        bins=[draft_edges, x_edges],
        weights=eligible["duration_hours"],
    )
    percent = histogram / denominator_hours * 100 if denominator_hours else histogram
    draft_labels = [_format_bin(value) for value in draft_edges[:-1]]
    x_labels = [_format_bin(value) for value in x_edges[:-1]]
    hours_df = pd.DataFrame(histogram, index=draft_labels, columns=x_labels)
    percent_df = pd.DataFrame(percent, index=draft_labels, columns=x_labels)
    table_hours = float(histogram.sum())
    coverage = table_hours / denominator_hours if denominator_hours else 0.0
    return ProfileResult(
        hours_df,
        percent_df,
        denominator_hours,
        denominator_hours,
        coverage,
        draft_edges,
        x_edges,
    )


def profile_with_totals(profile: pd.DataFrame) -> pd.DataFrame:
    if profile.empty:
        return profile.copy()
    result = profile.copy()
    result["Total"] = result.sum(axis=1)
    result.loc["Total"] = result.sum(axis=0)
    return result


def _month_floor(timestamp: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(timestamp.year, timestamp.month, 1)


def monthly_summary(segments: pd.DataFrame) -> pd.DataFrame:
    valid = segments[
        segments["segment_start"].notna()
        & segments["segment_end"].notna()
        & segments["duration_hours"].gt(0)
        & segments["duration_hours"].le(72)
    ].copy()
    if valid.empty:
        return pd.DataFrame()
    pieces: list[dict[str, Any]] = []
    for _, row in valid.iterrows():
        cursor = row["segment_start"]
        end = row["segment_end"]
        while cursor < end:
            month = _month_floor(cursor)
            next_month = month + pd.offsets.MonthBegin(1)
            piece_end = min(end, next_month)
            hours = (piece_end - cursor).total_seconds() / 3600
            pieces.append(
                {
                    "month": month,
                    "hours": hours,
                    "speed_hours": hours if 0 < row.get("speed_knots", np.nan) <= 40 else 0.0,
                    "speed_weighted": hours * row.get("speed_knots", np.nan)
                    if 0 < row.get("speed_knots", np.nan) <= 40
                    else 0.0,
                    "power_hours": hours if 0 < row.get("me_output_kw", np.nan) <= 100_000 else 0.0,
                    "power_weighted": hours * row.get("me_output_kw", np.nan)
                    if 0 < row.get("me_output_kw", np.nan) <= 100_000
                    else 0.0,
                    "temp_hours": hours if 0 < row.get("sea_temp_c", np.nan) <= 45 else 0.0,
                    "temp_weighted": hours * row.get("sea_temp_c", np.nan)
                    if 0 < row.get("sea_temp_c", np.nan) <= 45
                    else 0.0,
                }
            )
            cursor = piece_end
    split = pd.DataFrame(pieces)
    summary = split.groupby("month", as_index=False).sum(numeric_only=True)
    overall_start = valid["segment_start"].min()
    overall_end = valid["segment_end"].max()

    def available_hours(month: pd.Timestamp) -> float:
        start = max(overall_start, month)
        end = min(overall_end, month + pd.offsets.MonthBegin(1))
        return max((end - start).total_seconds() / 3600, 0.0)

    summary["available_hours"] = summary["month"].map(available_hours)
    summary["working_ratio_pct"] = np.where(
        summary["available_hours"] > 0,
        summary["hours"] / summary["available_hours"] * 100,
        np.nan,
    )
    for prefix, label in (
        ("speed", "avg_speed_knots"),
        ("power", "avg_me_output_kw"),
        ("temp", "avg_sea_temp_c"),
    ):
        summary[label] = np.where(
            summary[f"{prefix}_hours"] > 0,
            summary[f"{prefix}_weighted"] / summary[f"{prefix}_hours"],
            np.nan,
        )
    return summary[
        [
            "month", "hours", "available_hours", "working_ratio_pct",
            "avg_speed_knots", "avg_sea_temp_c", "avg_me_output_kw",
        ]
    ].rename(columns={"hours": "propelling_hours"})


def monthly_summary_excel(data_sum: pd.DataFrame) -> pd.DataFrame:
    """Calculate the 12-row Profile graph-source table from internal Data_sum."""
    valid = data_sum[data_sum["timestamp"].notna()].copy()
    if valid.empty:
        return pd.DataFrame()
    valid["month_start"] = valid["timestamp"].dt.to_period("M").dt.to_timestamp()
    rows: list[dict[str, Any]] = []
    for month, group in valid.groupby("month_start", sort=True):
        start = group["timestamp"].min()
        end = group["timestamp"].max()
        available = (end - start).total_seconds() / 3600
        propelling = float(group["duration_hours"].fillna(0).sum())
        rows.append(
            {
                "month": month,
                "data_start": start,
                "data_end": end,
                "available_hours": available,
                "propelling_hours": propelling,
                "working_ratio_pct": propelling / available * 100 if available > 0 else np.nan,
                "avg_speed_knots": group["speed_knots"].mean(),
                "avg_sea_temp_excel": group["data_sum_sea_temp"].mean(),
                "avg_actual_sea_temp_c": group["actual_sea_temp_c"].mean(),
                "avg_me_output_kw": group["me_output_kw"].mean(),
            }
        )
    # The supplied Profile sheet contains 12 monthly graph-source rows beginning
    # with the first Data_sum month. Graphs must use these rows directly.
    return pd.DataFrame(rows).head(12).reset_index(drop=True)


def excel_overall_summary(data_sum: pd.DataFrame) -> dict[str, Any]:
    """Calculate the Profile header directly from the internal Data_sum."""
    valid_times = data_sum["timestamp"].dropna()
    start = valid_times.min() if not valid_times.empty else pd.NaT
    end = valid_times.max() if not valid_times.empty else pd.NaT
    total_hours = (end - start).total_seconds() / 3600 if pd.notna(start) and pd.notna(end) else np.nan
    propelling_hours = float(data_sum["duration_hours"].fillna(0).sum())
    return {
        "year": start.year if pd.notna(start) else None,
        "data_start": start,
        "data_end": end,
        "total_hours": total_hours,
        "propelling_hours": propelling_hours,
        "working_ratio_pct": propelling_hours / total_hours * 100 if total_hours else np.nan,
        "avg_sea_temp_excel": data_sum["data_sum_sea_temp"].mean(),
        "avg_actual_sea_temp_c": data_sum["actual_sea_temp_c"].mean(),
        "avg_speed_knots": data_sum["speed_knots"].mean(),
    }


def fuel_consumption_summary(noon: ParsedReport, arrival: ParsedReport) -> dict[str, float]:
    def total(report: ParsedReport, field_name: str) -> float:
        if field_name not in report.data:
            return 0.0
        return float(pd.to_numeric(report.data[field_name], errors="coerce").fillna(0).sum())

    return {
        "noon_vlsfo": total(noon, "foc_vlsfo"),
        "noon_mgo": total(noon, "foc_mgo"),
        "arrival_vlsfo": total(arrival, "foc_vlsfo"),
        "arrival_mgo": total(arrival, "foc_mgo"),
    }


def sea_temperature_audit(data_sum: pd.DataFrame, noon: ParsedReport) -> dict[str, Any]:
    """Audit the Profile Sea Water Temp calculation from internal Data_sum.

    Internal Data_sum column G is populated from the physical Noon field identified
    by its two-row title, never from a fixed column position. The final average is
    calculated from Data_sum, with numeric zeroes included and blanks ignored.
    """
    source = noon.mapping.get("sea_temp")
    if source is None or "data_sum_sea_temp" not in data_sum:
        return {
            "valid": False,
            "source_column": "Internal Data_sum G",
            "source_header": "Sea Water Temp.",
            "upstream_column": None,
            "upstream_header": None,
            "numeric_count": 0,
            "zero_count": 0,
            "sum_value": np.nan,
            "average_value": np.nan,
            "minimum_value": np.nan,
            "maximum_value": np.nan,
            "out_of_range_count": 0,
            "zero_ratio": np.nan,
            "quality_warnings": [],
            "message": "Required title 'Sea Water temperature at noon' was not found.",
        }

    values = pd.to_numeric(data_sum["data_sum_sea_temp"], errors="coerce").dropna()
    header_tokens = set(source.normalized.split())
    source_matches = (
        ("seawater" in header_tokens or {"sea", "water"}.issubset(header_tokens))
        and bool({"temperature", "temp"} & header_tokens)
        and "noon" in header_tokens
        and not bool({"rolling", "period", "sec", "seconds"} & header_tokens)
    )
    zero_count = int(values.eq(0).sum()) if not values.empty else 0
    zero_ratio = zero_count / len(values) if len(values) else np.nan
    out_of_range_count = int((~values.between(-2, 40)).sum()) if not values.empty else 0
    quality_warnings: list[str] = []
    if len(values) and zero_ratio > 0.25:
        quality_warnings.append(
            f"{zero_count} of {len(values)} seawater-temperature readings are zero."
        )
    if out_of_range_count:
        quality_warnings.append(
            f"{out_of_range_count} seawater-temperature readings are outside -2 to 40 °C."
        )
    if values.empty:
        message = "Internal Data_sum Sea Water Temp contains no numeric readings."
    elif not source_matches:
        message = "The detected upstream title is not a physical seawater-temperature-at-noon field."
    else:
        message = (
            "Calculated from the physical seawater-temperature title, not a fixed Excel position. "
            f"Matched '{source.label}' in uploaded column {source.excel_column}."
        )
    return {
        "valid": bool(not values.empty and source_matches),
        "source_column": "Internal Data_sum G",
        "source_header": "Sea Water Temp.",
        "upstream_column": source.excel_column,
        "upstream_header": source.label,
        "numeric_count": int(values.count()),
        "zero_count": zero_count,
        "sum_value": float(values.sum()) if not values.empty else np.nan,
        "average_value": float(values.mean()) if not values.empty else np.nan,
        "minimum_value": float(values.min()) if not values.empty else np.nan,
        "maximum_value": float(values.max()) if not values.empty else np.nan,
        "out_of_range_count": out_of_range_count,
        "zero_ratio": zero_ratio,
        "quality_warnings": quality_warnings,
        "message": message,
    }


def mapping_table(report: ParsedReport) -> pd.DataFrame:
    rows = []
    for field_name, header in report.mapping.items():
        rows.append(
            {
                "Internal field": FIELD_TITLES.get(field_name, field_name),
                "Matched workbook header": header.label,
                "Excel column": header.excel_column,
            }
        )
    return pd.DataFrame(rows)


def vessel_values(report: ParsedReport) -> set[str]:
    if "vessel" not in report.data:
        return set()
    return {value for value in report.data["vessel"].map(_clean_key) if value}


def _vessel_identity_key(value: str) -> str:
    """Compare harmless punctuation/spacing variants without guessing vessel aliases."""
    return re.sub(r"[^A-Z0-9]+", "", _clean_key(value))


def validate_vessel_consistency(reports: dict[str, ParsedReport]) -> str:
    """Require exactly one vessel per report and the same vessel across all reports."""
    detected: dict[str, str] = {}
    for report_type in REPORT_NAMES:
        report = reports.get(report_type)
        if report is None:
            raise VesselValidationError(
                f"The {REPORT_NAMES[report_type]} report is unavailable for vessel validation."
            )
        values = sorted(vessel_values(report))
        if not values:
            raise VesselValidationError(
                f"The {REPORT_NAMES[report_type]} report contains no vessel name in the detected "
                "'Vessel' column."
            )
        if len(values) > 1:
            names = ", ".join(f"'{value}'" for value in values)
            raise VesselValidationError(
                f"The {REPORT_NAMES[report_type]} report contains multiple vessel names: {names}. "
                "Upload a report for one vessel only."
            )
        detected[report_type] = values[0]

    identity_keys = {_vessel_identity_key(value) for value in detected.values()}
    if len(identity_keys) != 1:
        details = "; ".join(
            f"{REPORT_NAMES[report_type]} = '{detected[report_type]}'"
            for report_type in REPORT_NAMES
        )
        raise VesselValidationError(
            f"Vessel mismatch: {details}. Upload Noon, Departure and Arrival reports for the same vessel."
        )
    return detected["noon"]
